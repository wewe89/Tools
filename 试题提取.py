#-*- coding:utf-8 -*-
import json
import openpyxl

fo=open(r"试题\data.txt","r",encoding='utf-8')
val=fo.read()
jsonstr=json.loads(val)
list=jsonstr["bizResult"]["examPaperItem"]

xlsfilename="试题\试题批量导入模板精简版.xlsx"
workbook=openpyxl.load_workbook(xlsfilename)
worksheet=workbook.worksheets[0]
index=2
for item in list:
    print(index)
    leixin=item["examItem"]["itemType"]
    itemType={"SINGLE":"单选","JUDGMENT":"判断","MULTIPLE":"多选"}.get(leixin)
    worksheet.cell(index,1,itemType)
    tigan=item["examItem"]["itemName"]
    worksheet.cell(index,2,tigan)
    itemOptions=item["examItem"]["itemOptions"]
    rightAnswer=item["rightAnswer"]
    daan=""
    print(itemOptions,leixin)
    #答案
    indexdaan=3
    if(leixin=="JUDGMENT"):
        print(type(rightAnswer),rightAnswer)
        if(rightAnswer=="true"):
            worksheet.cell(index, indexdaan, "正确")
        else:
            worksheet.cell(index, indexdaan, "错误")
    else:
        for option in itemOptions:
            worksheet.cell(index, indexdaan+option["showOrder"], option["content"])
            if option["itemOptionId"] in rightAnswer:
                daan=daan+","+{1: "A", 2: "B", 3: "C", 4: "C"}.get(option["showOrder"],"")
            worksheet.cell(index, indexdaan , daan[1:])
    index=index+1

workbook.save(xlsfilename)
print()